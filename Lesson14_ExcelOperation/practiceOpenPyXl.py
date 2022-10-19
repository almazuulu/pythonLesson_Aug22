from openpyxl import *
from openpyxl.styles import *

#Task 5
def taskFive():

    """
    Получите значение в диапазоне [‘A18:E30’] и назначьте ее в
    значение myListExcel
    """
    wb = load_workbook('food.xlsx')
    sheet = wb.worksheets[0]

    cells = sheet['A18':'E30']

    for row in cells:
        for v in row:
            print(v.value, end=' ')
        print()

def limitRows():
    wb = load_workbook('food.xlsx')
    sheet = wb.worksheets[0]
    nrows = 5

    datas = [ ]
    for i in range(2, nrows+2):
        cells = sheet[f'A{str(i)}:H{str(i)}']
        datas.append(cells[0])

    for row in datas:
        for v in row:
            print(v.value, end=' ')
        print()

def someWorkers(numRecord):
    wb2 = Workbook()
    sheet2 = wb2.worksheets[0]

    info = ['Имя', 'Возраст', 'Зарплата', 'Год работ']

    for i in range(len(info)):
        sheet2.cell(row=1, column=i+1).value = info[i]

    sheet2['A1'].font = Font(bold=True)
    sheet2['B1'].font = Font(bold=True)
    sheet2['C1'].font = Font(bold=True)

    allInfo = []
    for i in range(numRecord):
        name = input('Имя: ')
        age = int(input('Возраст: '))
        zp = int(input('Зарплата: '))
        ageJob = int(input('Год работ: '))
        allInfo.append([name,age,zp,ageJob])
    #[[name, age, zp, ageJob],[name, age, zp, ageJob]]
    print(allInfo)
    counter = 2

    for i in range(numRecord):
        sheet2.cell(row=counter, column=1).value = allInfo[i][0]
        sheet2.cell(row=counter, column=2).value = allInfo[i][1]
        sheet2.cell(row=counter, column=3).value = allInfo[i][2]
        sheet2.cell(row=counter, column=4).value = allInfo[i][3]
        counter += 1

    wb2.save('practice.xlsx')



if __name__=='__main__':
    # taskFive()
    #limitRows()
    numRecord = int(input('Сколько записей вы хотите сделать? '))
    someWorkers(numRecord)


