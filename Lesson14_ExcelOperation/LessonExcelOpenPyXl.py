# import openpyxl
from openpyxl import *
from openpyxl.styles import *

#Чтение данных из Excel
#Загрузка экзель и 1-ой страницы
wb = load_workbook('food.xlsx')
sheet = wb.worksheets[0] #1, 2, 3

#Чтение из определенных ячеек
cityName = sheet['C9'].value
print(cityName)

prodName = sheet['E14'].value
print(prodName)

print(type(cityName))
print(type(prodName))

infoProd = 'This ' + prodName + ' is from ' +  cityName
print(infoProd)

cells = sheet['C11': 'F11']
print(cells)

for el1, el2, el3, el4 in cells:
    print(el1.value, el2.value,el3.value, el4.value)

print('*'*100)
for el in cells[0]:
    print(el.value, end=' ')


cells2 = sheet['C12': 'G19']
print()
print('*'*20)

for row in cells2:
    for el in row:
        print(el.value, end =' ')
    print()


#2 Способ чтение из Экзель

dataSet1 = sheet.cell(row = 5, column=3).value
print(dataSet1)

dataSet2 =sheet.cell(row = 12, column=5).value
print(dataSet2)

#Чтение колонок
cityList = sheet['C'][1:]
for el in cityList:
    print(el.value, end=' ')

# prices = sheet['H']
# for el in prices:
#     print(el.value, end=' ')
print()
listUnitPrices = [float(el.value) for el in sheet['G'][1:]]
print(sum(listUnitPrices))

#Запись в Экзель при помощи OpenPyXL
countryList = ['France','USA', 'Kazakhstan', 'Uzgbekistan', 'Kyrgyzstan', 'China']
capitalList = ['Paris', 'Washington','Astana', 'Tashkent', 'Bishkek', 'Beijing']
vvpList = [5500, 8000, 4500, 3800, 2500, 10000]

wb2= Workbook()
sheet2 = wb2.worksheets[0]

cell1Header = sheet2.cell(row = 1, column = 1).value = 'Страна'
cell2Header = sheet2.cell(row = 1, column = 2).value = 'Столица'
cell3Header = sheet2.cell(row = 1, column = 3).value = 'ВВП на душу'

sheet2['A1'].font = Font(bold=True)
sheet2['B1'].font = Font(bold=True)
sheet2['C1'].font = Font(bold=True)

# sheet2.cell(row = 2, column= 1).value = countryList[0]
# sheet2.cell(row = 3, column= 1).value = countryList[1]
# sheet2.cell(row = 4, column= 1).value = countryList[2]

# sheet2.cell(row = 2, column= 2).value = capitalList[0]
# sheet2.cell(row = 3, column= 2).value = capitalList[1]
counter = 2
for i in range(6):
    sheet2.cell(row = counter, column = 1).value = countryList[i]
    sheet2.cell(row = counter, column = 2).value = capitalList[i]
    sheet2.cell(row = counter, column = 3).value = vvpList[i]
    counter += 1

wb2.save('countryList.xlsx')
print('Saved!')


















